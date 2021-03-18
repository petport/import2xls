from openpyxl import load_workbook
import csv
import numpy as np
import click
import string
import re
from xls2xlsx import XLS2XLSX


# --master_file /home/pedagil/PycharmProjects/xls_parser/dummy_xlsx.xlsx --input_file /home/pedagil/PycharmProjects/xls_parser/out_2.csv --master_ranges A17:B20_C21:D23 --input_file_ranges A1:B4_C7:D9

@click.command()
@click.option('--master_file', nargs=1,
              help='Master file. File to either select data to write to file, or import data from another file.')
@click.option('--output_file', nargs=1, help='Where to save the selected data.')
@click.option('--input_file', nargs=1, help='File from which to import data.')
@click.option('--master_ranges', nargs=1, help='One or more ranges. Use _ in between. Example: \"A1:B6_D1:E6\".')
@click.option('--input_file_ranges', nargs=1,
              help="Data to import from existing file. Use _ in between. Example: \"A1:B6_D1:E6\".")
@click.option('--convert_xls2xlsx', nargs=2, help='Convert .xlsx file in .xls format. ')
def main(master_file, output_file, input_file, master_ranges, input_file_ranges, convert_xls2xlsx):

    path_to_master_file_arg = master_file
    path_to_output_file_arg = output_file
    path_to_input_file_arg = input_file

    if convert_xls2xlsx is not None:
        x2x = XLS2XLSX(convert_xls2xlsx[0])
        x2x.to_xlsx(convert_xls2xlsx[1])

    if input_file_ranges is not None:
        input_file_ranges = input_file_ranges.split('_')
    master_ranges = master_ranges.split('_')

    if path_to_output_file_arg is not None and path_to_input_file_arg is None:
        batch_ranges = []

        if path_to_master_file_arg[-5:] == '.xlsx':
            wb = load_workbook(filename=path_to_master_file_arg)
            sheet = wb.worksheets[0]

            for item in master_ranges:
                rows = []
                for row in sheet.iter_rows(item):
                    rows.append([cell.value for cell in row])
                batch_ranges.append(np.array(rows))

            final_rows = np.hstack(batch_ranges)

        with open(path_to_output_file_arg, mode='w') as output_file:
            output_writer = csv.writer(output_file, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
            output_writer.writerows(final_rows)

    elif path_to_output_file_arg is None and path_to_input_file_arg is not None:
        if path_to_input_file_arg[-5:] == '.xlsx':
            batch_ranges = []
            wb_in = load_workbook(filename=path_to_input_file_arg)
            sheet_in = wb_in.worksheets[0]

            for item in input_file_ranges:
                rows = []
                for row in sheet_in.iter_rows(item):
                    rows.append([cell.value for cell in row])
                batch_ranges.append(np.array(rows))
            pass
        elif path_to_input_file_arg[-4:] == '.csv':

            temp_ranges = []
            for cell_range in input_file_ranges:
                temp_list = []
                for i in cell_range.split(':'):
                    temp = re.compile("([a-zA-Z]+)([0-9]+)")
                    res = temp.match(i).groups()
                    temp_list.append(list(res))
                temp_ranges.append(temp_list)

            with open(path_to_input_file_arg) as csv_file:
                csv_reader = csv.reader(csv_file, delimiter=',')

                data = []
                for row in csv_reader:
                    data.append(row)

                new_ranges = []
                for i in temp_ranges:
                    new_ranges.append([[col2num(pair[0]), int(pair[1])] for pair in i])

                batch_ranges = []

                for i in new_ranges:
                    temp_list = []
                    for row in data[i[0][1] - 1:i[1][1]]:
                        temp_list.append(row[i[0][0] - 1:i[1][0]])
                    batch_ranges.append(np.array(temp_list))

        # final_rows = np.hstack(batch_ranges)

        wb_master = load_workbook(filename=path_to_master_file_arg)
        sheet_out = wb_master.worksheets[0]

        flat_list = []
        for i in range(len(batch_ranges)):
            flat_list.append(list(batch_ranges[i].flatten()))

        for idx, item in enumerate(master_ranges):
            idx_dim2 = 0
            for row in sheet_out.iter_rows(item):
                for cell in row:
                    cell.value = flat_list[idx][idx_dim2]
                    idx_dim2 += 1

        wb_master.save(filename=path_to_master_file_arg)


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


if __name__ == '__main__':
    main()
